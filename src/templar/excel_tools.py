from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
from typing import NewType


Pathlib_Path = NewType('Pathlib_path', Path)


"""This module is intended for manipulating the data out of the excel file."""
# TODO: work out a way remove the empty cells from the returned data
# TODO: determine a way to extract the photo path from the cells.


def open_workbook(workbook_name: str,
                  directory_path: str = None) -> Pathlib_Path:
    """
    A function to allow the Excel spreadsheet to be easily opened. If the path is not provided or is not provided a string then use the CWD as the default directory.
    :param workbook_name:
    :param directory_path: path to the
    :return:
    """
    try:
        directory_path = Path(directory_path)
    except TypeError:
        directory_path = Path.cwd()
    finally:
        print("File directory: {}".format(directory_path))
        # pass
    workbook_path: Pathlib_Path = directory_path / str(workbook_name)
    print(workbook_path)
    return load_workbook(workbook_path)

def open_worksheet(worksheet_name, workbook_file):
    """
    A function to simplify the accessing of a worksheet in an already open workbook.
    :param worksheet_name:
    :param workbook_file:
    :return:
    """
    pass


dir = '/Users/tom/PycharmProjects/Autotemplate/autotemplate/'
file = 'test_excel.xlsm'

wb2 = open_workbook(file, dir)

print(wb2.sheetnames)
print(Path.cwd())
ws = wb2["Master List"]
# ws = wb2["Defect Lookups"]
for row in ws.rows:
    i = []
    for cell in row:
        i.append(cell.value)
    print(i)
    # for x, y in enumerate(i):
    #     print(x, y)
